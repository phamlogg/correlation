# coding: utf-8

# Import the required Libraries
from tkinter import *
from tkinter import ttk, filedialog
from tkinter.filedialog import askopenfile
import os
import platform
import sys
from functions import *
from unicodedata import name
import pandas as pd
import openpyxl as op
import os
import platform
import itertools
import time
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import ColumnDimension



filepath = None

def main():

    # Create an instance of tkinter frame
    win = Tk()

    # Set the geometry of tkinter frame
    win.geometry("700x350")

    def open_file():
        global filepath
        file = filedialog.askopenfile(mode='r', filetypes=[('Excel files', '*.xlsx')])
        if file:
            filepath = os.path.abspath(file.name)
            Label(win, text="The File is located at : " + str(filepath), font=('Aerial 11')).pack()

    # Add a Label widget
    label = Label(win, text="Click the Button to browse Excel File", font=('Georgia 13'))
    label.pack(pady=10)

    # Create a Button
    ttk.Button(win, text="Open Excel File", command=open_file).pack(pady=20)
    ttk.Button(win, text="Run", command=win.destroy).pack(pady=20)

    win.mainloop()

    delta = []

    wb1 = op.load_workbook(filepath)
    ws1 = wb1.active

    #List ra tên các tài sản & set cột delta
    try:
        name = list_assets(ws1)
    except Exception as e:
        print(e)
        exit(1)
    else:
        print('Liệt kê tài sản thành công!')

    if platform.system() == 'Windows':
        new_path = filepath.replace('\Data.xlsx', '\Output.xlsx')
    else:
        new_path = filepath.replace('/Data.xlsx', '/Output.xlsx')

    # Tính delta
    try:
        calc_delta(ws1)
    except Exception as e:
        print(e)
        exit(1)
    else:
        wb1.save(new_path)
        print('Tính delta thành công!')

    # Ghép tên theo cặp - nCr
    combinations = itertools.combinations(name,2)
    pair = list(combinations)
    for i in range(len(pair)):
        pair[i] = list(pair[i])
        pair[i] = pair[i][0] + "-" + pair[i][1] 

    # Tính standard dev.
    try:
        df = pd.read_excel(new_path)
        stdev_dict = calc_stdev(df, name, delta)
    except Exception as e:
        print(e)
        exit(1)
    else:
        print('Tính St. Dev thành công!')

    # Display stdev
    ws2 = wb1.create_sheet(title= "Output")

    titles = ["NAME", "ST.DEV", "PAIR", "-0.3 << 0.3", "PAIR", "-0.5 << 0.5"]

    for i, title in enumerate(titles):
        col_letter = op.utils.cell.get_column_letter(i+1)
        ws2.column_dimensions[col_letter] = ColumnDimension(ws2, bestFit=True)
        if title == 'PAIR':
            ws2.column_dimensions[col_letter].width = 18.0
        ws2.cell(row=1,column=i+1).value = title    

    i = 2
    for k, v in stdev_dict.items():
        cell = ws2.cell(row= i, column= 1)
        cell.value = str(k)
        cell = ws2.cell(row= i, column= 2)
        cell.value = round(float(v),5)
        i+=1

    corr = []

    #Tính tương quan 2 cột delta
    try:
        for i in range(0, len(delta)):
            for j in range(1, len(delta)):
                if i < j:
                    # print(delta[i])
                    # print(delta[j])
                    first_collumn = df[str(delta[i])]
                    second_collumn = df[str(delta[j])]
                    
                    correlation = first_collumn.corr(second_collumn)
                    
                    
                    # print(correlation)
                    corr.append(correlation)
    except Exception as e:
        print(e)
        exit(1)
    else:
        print('Tính correlation thành công!') 

    #Ghép hệ số tương quan với cặp tiền tương ứng
    try:
        correlation_dict = dict(zip(pair,corr))

        i = 2
        for k, v in correlation_dict.items():
            if round(float(v),5) <= 0.3 and round(float(v),5) >= -0.3:
                cell = ws2.cell(row= i, column= 3)
                cell.value = str(k)
                cell.font = Font(color = "00FF6600")
                cell = ws2.cell(row= i, column= 4)
                cell.value = round(float(v),5)
                cell.font = Font(color = "00FF6600")
                i+=1
            
        i = 2
        for k,v in correlation_dict.items():
            if (round(float(v),5) >= 0.3 and round(float(v),5) <= 0.5) or (round(float(v),5) <= -0.3 and round(float(v),5) >= -0.5):
                cell = ws2.cell(row= i, column= 5)
                cell.value = str(k)
                cell.font = Font(color = "00FF00FF")
                cell = ws2.cell(row= i, column= 6)
                cell.value = round(float(v),5)
                cell.font = Font(color = "00FF00FF")
                i+=1
            
    except Exception as e:
        print(e)
        exit(1)
    else:
        print('Ghép correlation thành công!')

    triplet = []

    for row in ws2.iter_rows(min_row=2, min_col=3, max_row=len(ws2['C']), max_col=3):
        for cell in row:
            for row in ws2.iter_rows(min_row=3, min_col=3, max_row=len(ws2['C']), max_col=3):
                for cell1 in row:
                    if cell.value != None and cell1.value != None:
                        if cell.value[:6] == cell1.value[:6]:
                            pair = cell.value[:6]
                            pair1 = cell.value[7:]
                            pair2 = cell1.value[7:]
                            try:
                                if -0.3 <= correlation_dict[f'{pair1}-{pair2}'] <= 0.3:
                                    triplet.append(f'{pair}-{pair1}-{pair2}')
                            except KeyError:
                                continue

    for index, triple in enumerate(triplet, 2):
        for row in ws2.iter_rows(min_row=int(index), min_col=7, max_row=len(triplet), max_col=7):
            for cell in row:
                cell.value = triplet[index - 2]
        
    wb1.save(new_path)

    if os.path.isfile(new_path) is True:
        print('Success!')
        time.sleep(5)
        sys.exit(0)
    else:
        print('Unsuccessful!')
        time.sleep(5)
        sys.exit(1)

main()



