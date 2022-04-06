#!/usr/bin/env python
# coding: utf-8

try:
    from functions import *
    from unicodedata import name
    import pandas as pd
    import openpyxl as op
    import os
    import itertools
    from sys import argv
    from openpyxl.styles import Font
except Exception as e:
    print(e)
    exit(1)
else:
    print('Import complete!')

delta = []


def calc_stdev(df):
    global name
    global delta
    std = []

    for i in range(len(df.columns)):
        if 'Delta' in df.columns[i]:
            delta.append(df.columns[i]) #Pick ra các cột số delta


    for i in range(0, len(delta)):
        column = df[str(delta[i])]
        stdev = column.std()
        std.append(stdev)

    stdev_dict = dict(zip(name,std))

    return stdev_dict
    

try:
    path = argv[1]
except Exception as e:
    print(e)
    exit(1)
else:
    print(path)
    print('File Found!')

wb1 = op.load_workbook(path)
ws1 = wb1.active

#List ra tên các tài sản & set cột delta
try:
    name = list_assets(ws1)
except Exception as e:
    print(e)
    exit(1)
else:
    print('Liệt kê tài sản thành công!')

new_path = path.replace('/Data.xlsx', '/Output.xlsx')

# Tính delta
try:
    calc_delta(ws1)
except Exception as e:
    print(e)
    exit(1)
else:
    wb1.save(new_path)
    print('Tính delta thành công!')

# Ghép tên theo cặp - nCr
combinations = itertools.combinations(name,2)
pair = list(combinations)
for i in range(len(pair)):
    pair[i] = list(pair[i])
    pair[i] = pair[i][0] + "-" + pair[i][1] 

# Tính standard dev.
try:
    df = pd.read_excel(new_path)
    stdev_dict = calc_stdev(df)
except Exception as e:
    print(e)
    exit(1)
else:
    print('Tính St. Dev thành công!')

# Display stdev
ws2 = wb1.create_sheet(title= "Output")

title = ["Name", "Standard Dev.", "Pair", "-0.3 << 0.3", "Pair", "-0.5 << 0.5"]

for i in range(1,7):
    cell = ws2.cell(row=1,column=i)
    cell.value = title[i - 1]

i = 2
for k, v in stdev_dict.items():
    cell = ws2.cell(row= i, column= 1)
    cell.value = str(k)
    cell = ws2.cell(row= i, column= 2)
    cell.value = round(float(v),5)
    i+=1

corr = []

#Tính tương quan 2 cột delta
try:
    for i in range(0, len(delta)):
        for j in range(1, len(delta)):
            if i < j:
                # print(delta[i])
                # print(delta[j])
                first_collumn = df[str(delta[i])]
                second_collumn = df[str(delta[j])]
                
                correlation = first_collumn.corr(second_collumn)
                
                
                # print(correlation)
                corr.append(correlation)
except Exception as e:
    print(e)
    exit(1)
else:
    print('Tính correlation thành công!') 

#Ghép hệ số tương quan với cặp tiền tương ứng
try:
    correlation_dict = dict(zip(pair,corr)) 

    i = 2
    for k, v in correlation_dict.items():
        if round(float(v),5) <= 0.3 and round(float(v),5) >= -0.3:
            cell = ws2.cell(row= i, column= 3)
            cell.value = str(k)
            cell.font = Font(color = "00FF6600")
            cell = ws2.cell(row= i, column= 4)
            cell.value = round(float(v),5)
            cell.font = Font(color = "00FF6600")
            i+=1
    
    i = 2
    for k,v in correlation_dict.items():
        if (round(float(v),5) >= 0.3 and round(float(v),5) <= 0.5) or (round(float(v),5) <= -0.3 and round(float(v),5) >= -0.5):
            cell = ws2.cell(row= i, column= 5)
            cell.value = str(k)
            cell.font = Font(color = "00FF00FF")
            cell = ws2.cell(row= i, column= 6)
            cell.value = round(float(v),5)
            cell.font = Font(color = "00FF00FF")
            i+=1
        
except Exception as e:
    print(e)
    exit(1)
else:
    print('Ghép correlation thành công!')

wb1.save(new_path)

if os.path.isfile(new_path) is True:
    print('Success!')
    exit(0)
else:
    print('Unsuccessful!')
    exit(1)